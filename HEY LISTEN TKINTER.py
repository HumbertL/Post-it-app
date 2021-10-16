import tkinter.messagebox
import openpyxl
#Marbeta gewidmet, danke für deine Geduld und Freundlichkeit mit mir wenn ich verärgert war. Ohne dich hätte ich dieses Projekt nicht zu Ende erreichen können.
wb = openpyxl.load_workbook('POST IT APP.xlsx') #ABRE EL DOCUMENTO
root = tkinter.Tk()
root.title("Post its")
root.configure(background='#e3ac40') #Cambiar de color la ventana principal, aunque ni siquiera se ve
#como las funciones de abajo indican cada una es para abrir otra subventana el proceso por cada subventana es el mismo
def openlunes():
    wb = openpyxl.load_workbook('POST IT APP.xlsx')
    wb.active= 0
    sheet = wb.active
    #defino los comandos asignaré a los botones de esta ventana
    def add_task():
        #comando del boton agregar tareas
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 0
        sheet = wb.active
        
        task = entry_task.get()
        SiguienteInsert = sheet.max_row
        if task != "":
            if sheet['B2'].value == None:
                sheet.cell(row=2, column=2).value = task
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                wb.save('POST IT APP.xlsx')
                wb.close()
            else:
                postithomie = []
                postithomie.append(SiguienteInsert)
                postithomie.append(task)
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                sheet.append(postithomie)
                wb.save('POST IT APP.xlsx')
                wb.close()
        else:
            tkinter.messagebox.showwarning(title="Warning!", message="Por favor escribe algo")
    def delete_task(): #comando del boton borrar tarea 
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 0
        sheet = wb.active
        try:
            task_index = listbox_tasks.curselection()[0]
            listbox_tasks.delete(task_index) #index o mouse del usuario cuando selecciona una tarea 
            sheet.delete_rows(task_index + 2, 2) #deletea en base al index dentro de la hoja del documento
        except:
            tkinter.messagebox.showwarning(title="Achtung!", message="Por favor selecciona una tarea")
        finally:
            wb.save('POST IT APP.xlsx')
    def modificartarea():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 0
        sheet = wb.active
        task = entry_task.get()
        task_index = listbox_tasks.curselection()[0]
        if task != "":
            listbox_tasks.delete(task_index)
            listbox_tasks.insert(task_index, task)
            sheet.cell(row=task_index + 2, column=2).value = task
            wb.save('POST IT APP.xlsx')
            entry_task.delete(0, tkinter.END)
            wb.close()
        else:
            tkinter.messagebox.showwarning(title="Advertencia", message="Por favor escribe algo")

    #creando la ventana
    lunes = tkinter.Toplevel(root) #dice que abrira la ventana a partir de la anterior
    lunes.title("lunes") #titulo de la ventana
    lunes.configure(background='#f55a42') #color de la ventana
    frame_tasks = tkinter.Frame(lunes)
    frame_tasks.pack()
    listbox_tasks = tkinter.Listbox(frame_tasks, height=10, width=50, background='#f55a42' ) #lista de tareas donde se va a ver lo de excel
    for r in range(2, sheet.max_row + 1): #muestra solo las celdas del documento sin los headers de "no. de tarea y tarea"
        for c in range(1, sheet.max_column + 1):
            celda = (sheet.cell(row=r, column=c).value)
        listbox_tasks.insert(tkinter.END, celda)
    listbox_tasks.pack(side=tkinter.LEFT)
    scrollbar_tasks = tkinter.Scrollbar(frame_tasks)
    scrollbar_tasks.pack(side=tkinter.RIGHT, fill=tkinter.Y)

    listbox_tasks.config(yscrollcommand=scrollbar_tasks.set, background = '#b54331') #scrollbars por si hay muchisimas tareas
    scrollbar_tasks.config(command=listbox_tasks.yview, background = '#b54331')

    entry_task = tkinter.Entry(lunes, width=50) #cuadro donde el usuario puede escribir para cambiar/añadir una tarea
    entry_task.config(background = '#cc6454')
    entry_task.pack()
    button_add_task = tkinter.Button(lunes, text="Añadir tarea", width=48, command=add_task, background = '#b54331')
    button_add_task.pack()
    button_delete_task = tkinter.Button(lunes, text="Borrar tarea", width=48, command=delete_task, background = '#b54331')
    button_delete_task.pack()
    button_edit_tarea =  tkinter.Button(lunes, text="Modificar tarea", width=48, command=modificartarea, background = '#b54331')
    button_edit_tarea.pack()
def openmartes():
    wb = openpyxl.load_workbook('POST IT APP.xlsx')
    wb.active = 1
    sheet = wb.active
    

    def add_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 1
        sheet = wb.active
        
        task = entry_task.get()
        SiguienteInsert = sheet.max_row
        if task != "":
            if sheet['B2'].value == None:
                sheet.cell(row=2, column=2).value = task
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                wb.save('POST IT APP.xlsx')
                wb.close()
            else:
                postithomie = []
                postithomie.append(SiguienteInsert)
                postithomie.append(task)
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                sheet.append(postithomie)
                wb.save('POST IT APP.xlsx')
        else:
            tkinter.messagebox.showwarning(title="Warning!", message="Please enter a task, homie")

    def delete_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 1
        sheet = wb.active
        try:
            task_index = listbox_tasks.curselection()[0]
            listbox_tasks.delete(task_index)
            sheet.delete_rows(task_index + 2, 2)
        except:
            tkinter.messagebox.showwarning(title="Achtung!", message="Por favor selecciona una tarea")
        finally:
            wb.save('POST IT APP.xlsx')
            wb.close()

    def modificartarea():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 1
        sheet = wb.active
        task = entry_task.get()
        task_index = listbox_tasks.curselection()[0]
        if task != "":
            listbox_tasks.delete(task_index)
            listbox_tasks.insert(task_index, task)
            sheet.cell(row=task_index + 2, column=2).value = task
            wb.save('POST IT APP.xlsx')
            entry_task.delete(0, tkinter.END)
            wb.close()
        else:
            tkinter.messagebox.showwarning(title="Advertencia", message="Por favor escribe algo")
    martes= tkinter.Toplevel(root)
    martes.title("martes")
    martes.configure(background='#fc9003')
    frame_tasks = tkinter.Frame(martes)
    frame_tasks.pack()
    listbox_tasks = tkinter.Listbox(frame_tasks, height=10, width=50, background='#fc9003')
    for r in range(2, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            celda = sheet.cell(row=r, column=c).value
        listbox_tasks.insert(tkinter.END, celda)
    listbox_tasks.pack(side=tkinter.LEFT)
    scrollbar_tasks = tkinter.Scrollbar(frame_tasks)
    scrollbar_tasks.pack(side=tkinter.RIGHT, fill=tkinter.Y)

    listbox_tasks.config(yscrollcommand=scrollbar_tasks.set, background = '#d97f0b')
    scrollbar_tasks.config(command=listbox_tasks.yview, background = '#d97f0b')

    entry_task = tkinter.Entry(martes, width=50)
    entry_task.config(background = '#f09b2e')
    entry_task.pack()
    button_add_task = tkinter.Button(martes, text="Añadir tarea", width=48, command=add_task, background = '#ba6d09')
    button_add_task.pack()

    button_delete_task = tkinter.Button(martes, text="Borrar tarea", width=48, command=delete_task, background = '#ba6d09')
    button_delete_task.pack()

    button_edit_tarea = tkinter.Button(martes, text="Modificar tarea", width=48, command=modificartarea, background = '#ba6d09')
    button_edit_tarea.pack()
def openmiercoles():
    wb = openpyxl.load_workbook('POST IT APP.xlsx')
    wb.active = 2
    sheet = wb.active
    

    def add_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 2
        sheet = wb.active
        
        task = entry_task.get()
        SiguienteInsert = sheet.max_row
        if task != "":
            if sheet['B2'].value == None:
                sheet.cell(row=2, column=2).value = task
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                wb.save('POST IT APP.xlsx')
                wb.close()
            else:
                postithomie = []
                postithomie.append(SiguienteInsert)
                postithomie.append(task)
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                sheet.append(postithomie)
                wb.save('POST IT APP.xlsx')
                wb.close()
        else:
            tkinter.messagebox.showwarning(title="Warning!", message="Please enter a task, homie")

    def delete_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 2
        sheet = wb.active
        try:
            task_index = listbox_tasks.curselection()[0]
            listbox_tasks.delete(task_index)
            sheet.delete_rows(task_index + 2, 2)
        except:
            tkinter.messagebox.showwarning(title="Achtung!", message="Por favor selecciona una tarea")
        finally:
            wb.save('POST IT APP.xlsx')
            wb.close()

    def modificartarea():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 3
        sheet = wb.active
        task = entry_task.get()
        task_index = listbox_tasks.curselection()[0]
        if task != "":
            listbox_tasks.delete(task_index)
            listbox_tasks.insert(task_index, task)
            sheet.cell(row=task_index + 2, column=2).value = task
            wb.save('POST IT APP.xlsx')
            entry_task.delete(0, tkinter.END)
            wb.close()
        else:
            tkinter.messagebox.showwarning(title="Advertencia", message="Por favor escribe algo")
    miercoles= tkinter.Toplevel(root)
    miercoles.title("miercoles")
    miercoles.configure(background='#3bdbc6')
    frame_tasks = tkinter.Frame(miercoles)
    frame_tasks.pack()
    listbox_tasks = tkinter.Listbox(frame_tasks, height=10, width=50, background='#3bdbc6')
    list = ''
    for r in range(2, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            celda = (sheet.cell(row=r, column=c).value)
        listbox_tasks.insert(tkinter.END, celda)
    listbox_tasks.pack(side=tkinter.LEFT)
    scrollbar_tasks = tkinter.Scrollbar(frame_tasks)
    scrollbar_tasks.pack(side=tkinter.RIGHT, fill=tkinter.Y)

    listbox_tasks.config(yscrollcommand=scrollbar_tasks.set, background = '#2a9c8d')
    scrollbar_tasks.config(command=listbox_tasks.yview, background = '#2a9c8d')

    entry_task = tkinter.Entry(miercoles, width=50)
    entry_task.config(background = '#59e3d1')
    entry_task.pack()
    button_add_task = tkinter.Button(miercoles, text="Añadir tarea", width=48, command=add_task, background = '#2a9c8d')
    button_add_task.pack()

    button_delete_task = tkinter.Button(miercoles, text="Borrar tarea", width=48, command=delete_task, background = '#2a9c8d')
    button_delete_task.pack()

    button_edit_tarea = tkinter.Button(miercoles, text="Modificar tarea", width=48, command=modificartarea, background = '#2a9c8d')
    button_edit_tarea.pack()
def openjueves():
    wb = openpyxl.load_workbook('POST IT APP.xlsx')
    wb.active = 3
    sheet = wb.active
    

    def add_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 3
        sheet = wb.active
        
        task = entry_task.get()
        SiguienteInsert = sheet.max_row
        if task != "":
            if sheet['B2'].value == None:
                sheet.cell(row=2, column=2).value = task
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                wb.save('POST IT APP.xlsx')
                wb.close()
            else:
                postithomie = []
                postithomie.append(SiguienteInsert)
                postithomie.append(task)
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                sheet.append(postithomie)
                wb.save('POST IT APP.xlsx')
                wb.close()
        else:
            tkinter.messagebox.showwarning(title="Warning!", message="Please enter a task, homie")

    def delete_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 3
        sheet = wb.active
        try:
            task_index = listbox_tasks.curselection()[0]
            listbox_tasks.delete(task_index)
            sheet.delete_rows(task_index + 2, 2)
        except:
            tkinter.messagebox.showwarning(title="Achtung!", message="Por favor selecciona una tarea")
        finally:
            wb.save('POST IT APP.xlsx')
            wb.close()

    def modificartarea():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 3
        sheet = wb.active
        task = entry_task.get()
        task_index = listbox_tasks.curselection()[0]
        if task != "":
            listbox_tasks.delete(task_index)
            listbox_tasks.insert(task_index, task)
            sheet.cell(row=task_index + 2, column=2).value = task
            wb.save('POST IT APP.xlsx')
            entry_task.delete(0, tkinter.END)
            wb.close()
        else:
            tkinter.messagebox.showwarning(title="Advertencia", message="Por favor escribe algo")
    jueves = tkinter.Toplevel(root)
    jueves.title("jueves")
    jueves.configure(background='#f02252')
    frame_tasks = tkinter.Frame(jueves)
    frame_tasks.pack()
    listbox_tasks = tkinter.Listbox(frame_tasks, height=10, width=50, background='#f02252')
    for r in range(2, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            celda = (sheet.cell(row=r, column=c).value)
        listbox_tasks.insert(tkinter.END, celda)
    listbox_tasks.pack(side=tkinter.LEFT)
    scrollbar_tasks = tkinter.Scrollbar(frame_tasks)
    scrollbar_tasks.pack(side=tkinter.RIGHT, fill=tkinter.Y)

    listbox_tasks.config(yscrollcommand=scrollbar_tasks.set, background='#f02252')
    scrollbar_tasks.config(command=listbox_tasks.yview, background='#f02252')

    entry_task = tkinter.Entry(jueves, width=50)
    entry_task.config(background='#ab3c55')
    entry_task.pack()
    button_add_task = tkinter.Button(jueves, text="Añadir tarea", width=48, command=add_task, background='#9c1937')
    button_add_task.pack()

    button_delete_task = tkinter.Button(jueves, text="Borrar tarea", width=48, command=delete_task, background='#9c1937')
    button_delete_task.pack()

    button_edit_tarea = tkinter.Button(jueves, text="Modificar tarea", width=48, command=modificartarea, background='#9c1937')
    button_edit_tarea.pack()
def openviernes():
    wb = openpyxl.load_workbook('POST IT APP.xlsx')
    wb.active = 4
    sheet = wb.active
    

    def add_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 4
        sheet = wb.active
        
        task = entry_task.get()
        SiguienteInsert = sheet.max_row
        if task != "":
            if sheet['B2'].value == None:
                sheet.cell(row=2, column=2).value = task
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                wb.save('POST IT APP.xlsx')
                wb.close()
            else:
                postithomie = []
                postithomie.append(SiguienteInsert)
                postithomie.append(task)
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                sheet.append(postithomie)
                wb.save('POST IT APP.xlsx')
                wb.close()
        else:
            tkinter.messagebox.showwarning(title="Warning!", message="Please enter a task, homie")

    def delete_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 4
        sheet = wb.active
        try:
            task_index = listbox_tasks.curselection()[0]
            listbox_tasks.delete(task_index)
            sheet.delete_rows(task_index + 2, 2)
        except:
            tkinter.messagebox.showwarning(title="Achtung!", message="Por favor selecciona una tarea")
        finally:
            wb.save('POST IT APP.xlsx')
            wb.close()

    def modificartarea():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 4
        sheet = wb.active
        task = entry_task.get()
        task_index = listbox_tasks.curselection()[0]
        if task != "":
            listbox_tasks.delete(task_index)
            listbox_tasks.insert(task_index, task)
            sheet.cell(row=task_index + 2, column=2).value = task
            wb.save('POST IT APP.xlsx')
            entry_task.delete(0, tkinter.END)
            wb.close()
        else:
            tkinter.messagebox.showwarning(title="Advertencia", message="Por favor escribe algo")
    viernes = tkinter.Toplevel(root)
    viernes.title("viernes")
    viernes.configure(background='#60e354')
    frame_tasks = tkinter.Frame(viernes)
    frame_tasks.pack()
    listbox_tasks = tkinter.Listbox(frame_tasks, height=10, width=50, background='#60e354')
    for r in range(2, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            celda = (sheet.cell(row=r, column=c).value)
        listbox_tasks.insert(tkinter.END, celda)
    listbox_tasks.pack(side=tkinter.LEFT)
    scrollbar_tasks = tkinter.Scrollbar(frame_tasks)
    scrollbar_tasks.pack(side=tkinter.RIGHT, fill=tkinter.Y)

    listbox_tasks.config(yscrollcommand=scrollbar_tasks.set, background='#60e354')
    scrollbar_tasks.config(command=listbox_tasks.yview, background='#60e354')

    entry_task = tkinter.Entry(viernes, width=50)
    entry_task.config(background='#86e87d')
    entry_task.pack()
    button_add_task = tkinter.Button(viernes, text="Añadir tarea", width=48, command=add_task, background='#45a63c')
    button_add_task.pack()

    button_delete_task = tkinter.Button(viernes, text="Borrar tarea", width=48, command=delete_task, background='#45a63c')
    button_delete_task.pack()

    button_edit_tarea = tkinter.Button(viernes, text="Modificar tarea", width=48, command=modificartarea, background='#45a63c')
    button_edit_tarea.pack()
def opensabado():
    wb = openpyxl.load_workbook('POST IT APP.xlsx')
    wb.active = 5
    sheet = wb.active
    

    def add_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 5
        sheet = wb.active
        task = entry_task.get()
        SiguienteInsert = sheet.max_row
        if task != "":
            if sheet['B2'].value == None:
                sheet.cell(row=2, column=2).value = task
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                wb.save('POST IT APP.xlsx')
                wb.close()
            else:
                postithomie = []
                postithomie.append(SiguienteInsert)
                postithomie.append(task)
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                sheet.append(postithomie)
                wb.save('POST IT APP.xlsx')
                wb.close()
        else:
            tkinter.messagebox.showwarning(title="Warning!", message="Please enter a task, homie")

    def delete_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 5
        sheet = wb.active
        try:
            task_index = listbox_tasks.curselection()[0]
            listbox_tasks.delete(task_index)
            sheet.delete_rows(task_index + 2, 2)
        except:
            tkinter.messagebox.showwarning(title="Achtung!", message="Por favor selecciona una tarea")
        finally:
            wb.save('POST IT APP.xlsx')
            wb.close()
    def modificartarea():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 5
        sheet = wb.active
        task = entry_task.get()
        task_index = listbox_tasks.curselection()[0]
        task = int(task)
        if task != "":
            listbox_tasks.delete(task_index)
            listbox_tasks.insert(task_index, task)
            sheet.cell(row=task_index + 2, column=2).value = task
            wb.save('POST IT APP.xlsx')
            entry_task.delete(0, tkinter.END)
            wb.close()
        else:
            tkinter.messagebox.showwarning(title="Advertencia", message="Por favor escribe algo")
    sabado = tkinter.Toplevel(root)
    sabado.title("sabado")
    sabado.configure(background='#9883a4')
    frame_tasks = tkinter.Frame(sabado)
    frame_tasks.pack()
    listbox_tasks = tkinter.Listbox(frame_tasks, height=10, width=50, background='#9883a4')
    for r in range(2, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            celda = (sheet.cell(row=r, column=c).value)
        listbox_tasks.insert(tkinter.END, celda)
    listbox_tasks.pack(side=tkinter.LEFT)
    scrollbar_tasks = tkinter.Scrollbar(frame_tasks)
    scrollbar_tasks.pack(side=tkinter.RIGHT, fill=tkinter.Y)

    listbox_tasks.config(yscrollcommand=scrollbar_tasks.set, background='#806e8a')
    scrollbar_tasks.config(command=listbox_tasks.yview, background='#806e8a')

    entry_task = tkinter.Entry(sabado, width=50)
    entry_task.config(background='#c8b4d4')
    entry_task.pack()
    button_add_task = tkinter.Button(sabado, text="Añadir tarea", width=48, command=add_task, background='#72637a')
    button_add_task.pack()

    button_delete_task = tkinter.Button(sabado, text="Borrar tarea", width=48, command=delete_task, background='#72637a')
    button_delete_task.pack()

    button_edit_tarea = tkinter.Button(sabado, text="Modificar tarea", width=48, command=modificartarea, background='#72637a')
    button_edit_tarea.pack()
def opendomingo():
    wb = openpyxl.load_workbook('POST IT APP.xlsx')
    wb.active = 6
    sheet = wb.active
    

    def add_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 6
        sheet = wb.active
        
        task = entry_task.get()
        SiguienteInsert = sheet.max_row
        if task != "":
            if sheet['B2'].value == None:
                sheet.cell(row=2, column=2).value = task
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                wb.save('POST IT APP.xlsx')
                wb.close()
            else:
                postithomie = []
                postithomie.append(SiguienteInsert)
                postithomie.append(task)
                listbox_tasks.insert(tkinter.END, task)
                entry_task.delete(0, tkinter.END)
                sheet.append(postithomie)
                wb.save('POST IT APP.xlsx')
                wb.close()
        else:
            tkinter.messagebox.showwarning(title="Warning!", message="Please enter a task, homie")

    def delete_task():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 6
        sheet = wb.active
        try:
            task_index = listbox_tasks.curselection()[0]
            listbox_tasks.delete(task_index)
            sheet.delete_rows(task_index + 2, 2)
        except:
            tkinter.messagebox.showwarning(title="Achtung!", message="Por favor selecciona una tarea")
        finally:
            wb.save('POST IT APP.xlsx')
            wb.close()

    def modificartarea():
        wb = openpyxl.load_workbook('POST IT APP.xlsx')
        wb.active = 6
        sheet = wb.active
        task = entry_task.get()
        task_index = listbox_tasks.curselection()[0]
        if task != "":
            listbox_tasks.delete(task_index)
            listbox_tasks.insert(task_index, task)
            sheet.cell(row=task_index + 2, column=2).value = task
            wb.save('POST IT APP.xlsx')
            entry_task.delete(0, tkinter.END)
            wb.close()
        else:
            tkinter.messagebox.showwarning(title="Advertencia", message="Por favor escribe algo")
    domingo = tkinter.Toplevel(root)
    domingo.title("domingo")
    domingo.configure(background='#04889d')
    frame_tasks = tkinter.Frame(domingo)
    frame_tasks.pack()
    listbox_tasks = tkinter.Listbox(frame_tasks, height=10, width=50, background='#04889d')
    for r in range(2, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            celda = (sheet.cell(row=r, column=c).value)
        listbox_tasks.insert(tkinter.END, celda)
    listbox_tasks.pack(side=tkinter.LEFT)
    scrollbar_tasks = tkinter.Scrollbar(frame_tasks)
    scrollbar_tasks.pack(side=tkinter.RIGHT, fill=tkinter.Y)

    listbox_tasks.config(yscrollcommand=scrollbar_tasks.set, background='#04889d')
    scrollbar_tasks.config(command=listbox_tasks.yview, background='#04889d')

    entry_task = tkinter.Entry(domingo, width=50)
    entry_task.config(background='#23acc2')
    entry_task.pack()
    button_add_task = tkinter.Button(domingo, text="Añadir tarea", width=48, command=add_task, background='#026a7a')
    button_add_task.pack()

    button_delete_task = tkinter.Button(domingo, text="Borrar tarea", width=48, command=delete_task, background='#026a7a')
    button_delete_task.pack()
    button_edit_tarea = tkinter.Button(domingo, text="Modificar tarea", width=48, command=modificartarea, background='#026a7a')
    button_edit_tarea.pack()
# muestra los botones con colores para cada dia 
lunesbutton = tkinter.Button(root, text="lunes", width=48, font = ('yu gothic', 10),bg='#f55a42', command=openlunes)
lunesbutton.pack()
martesbutton = tkinter.Button(root, text="martes", width=48,  font = ('yu gothic', 10), bg='#fc9003', command=openmartes)
martesbutton.pack()
miercolesbutton = tkinter.Button(root, text="miercoles",  width=48, font = ('yu gothic', 10), bg='#3bdbc6', command=openmiercoles)
miercolesbutton.pack()
juevesbutton = tkinter.Button(root, text="jueves", width=48, font = ('yu gothic', 10), bg='#f02252', command=openjueves)
juevesbutton.pack()
viernesbutton = tkinter.Button(root, text="viernes", width=48, font = ('yu gothic', 10), bg='#60e354', command=openviernes)
viernesbutton.pack()
sabadobutton = tkinter.Button(root, text="sabado", width=48, font = ('yu gothic', 10), bg='#9883a4', command=opensabado)
sabadobutton.pack()
domingobutton = tkinter.Button(root, text="domingo", width=48, font = ('yu gothic', 10), bg='#04889d', command=opendomingo)
domingobutton.pack()
root.mainloop()