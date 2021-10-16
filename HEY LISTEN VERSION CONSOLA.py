from datetime import datetime
import openpyxl
def addtasks(ws):
    SiguienteInsert = ws.max_row #renglón por el cual vamos a empezar agregar tareas
    path = "POST IT APP.xlsx"
    wb = openpyxl.load_workbook(path)
    SiguienteInsert=ws.max_row
    number = 0
    taskscell = ws['B2']
    #este es solamente los parametros a seguir para mostrar las tareas
    for r in range(2, ws.max_row + 1):
        number = number + 1
        for c in range(1, ws.max_column):
            ws.cell(row=r, column=c).value = number
    for i in range(1, ws.max_column + 1):
        print(ws.cell(row=1, column=i).value, end='\t')
    print('\n')
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            print(ws.cell(row=r, column=c).value, end='\t\t')
        print()
    print(colored("\n Si quieres ingresar más tareas escribe si", "blue"))
    print(colored("\n Si ingresaras más tareas, escribe stop cuando termines", "blue"))
    print(colored("\nSi no necesitas escribir más tareas escribe no", "blue"))
    yea=input(colored("\n Estas segur@ que quieres agregar más tareas?", "blue")) #confirmación para entrar al ciclo
    if yea=="no" or yea == "NO" or yea == "No":
        print("\n Que bueno que no tengas más tareas, disfruta tu día ")
    while yea=="si" or yea == "Si" or yea == "SI": #entra al ciclo hasta que el usuario escribe "STOP" o "stop"
           tarea = input(colored("\n Escribe la tarea que quieres agregar: ", "blue"))
           while tarea == "" or tarea == " ":
               tarea = input(colored("\n Por favor escribe algo", "blue"))
            #este siguiente if es para verificar si la hoja esta vacía ya que tuvimos un problema con el indice de tareas cuando la hoja esta vacía
           if taskscell.value == None:
               ws.cell(row=2, column=2).value = tarea
           else:
               postithomie=[] #lista vacia utilizada durante el ciclo/función
               postithomie.append(SiguienteInsert) #escribe en la lista vacía todas las tareas de la hoja
               postithomie.append(tarea) #escribe en la ultima posicion de la lista generada
               if tarea=="stop" or tarea=="STOP" or tarea == "stop":
                   break
               try:
                   ws.append(postithomie) #inserta dicha lista en la hoja
                   wb.save("Post IT APP.xlsx") #guarda el documento
                   print(colored("Insertado Exitosamente", "blue"))
               except:
                   print(colored("Ocurrió un error al escribir ", "blue"))
               finally:
                   wb.close()
def menu(tag):
    #Este es el menu que mostrara las tareas por día del usuario para eso es la librería de datetime
    wb = openpyxl.load_workbook('POST IT APP.xlsx')
    if tag == "Monday":
        #Si es lunes mostrará las tareas del lunes osea la hoja "0" en excel
        print("Lunes")
        wb.active = 0
        sheet = wb.active
        print("\nImprime las cabeceras de la hoja de cálculo")
        for i in range(1, sheet.max_column + 1):
            print(sheet.cell(row=1, column=i).value, end='\t')
        print('\n')
        for r in range(2, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                print(sheet.cell(row=r, column=c).value, end='\t\t')
            print()
    if tag == "Tuesday":
        # Si es martes mostrará las tareas del martes osea la hoja "1" en excel
        print("Martes")
        wb.active = 1
        sheet = wb.active
        for i in range(1, sheet.max_column + 1):
            print(sheet.cell(row=1, column=i).value, end='\t')
        print('\n')
        for r in range(2, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                print(sheet.cell(row=r, column=c).value, end='\t\t')
            print()
    if tag == "Wednesday":
        # Si es Miercoles  mostrará las tareas del martes osea la hoja "2" en excel
        print("Miercoles")
        wb.active = 2
        sheet = wb.active
        for i in range(1, sheet.max_column + 1):
            print(sheet.cell(row=1, column=i).value, end='\t')
        print('\n')
        for r in range(2, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                print(sheet.cell(row=r, column=c).value, end='\t\t')
            print()
    if tag == "Thursday":
        print("Jueves")
        wb.active = 3
        sheet = wb.active
        for i in range(1, sheet.max_column + 1):
            print(sheet.cell(row=1, column=i).value, end='\t')
        print('\n')
        for r in range(2, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                print(sheet.cell(row=r, column=c).value, end='\t\t')
            print()
    if tag == "Friday":
        print("Viernes")
        wb.active = 4
        sheet = wb.active
        #imprime cabeceras
        for i in range(1, sheet.max_column + 1):
            print(sheet.cell(row=1, column=i).value, end='\t')
        print('\n')
        #imprime las tareas de ese día
        for r in range(2, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                print(sheet.cell(row=r, column=c).value, end='\t\t')
            print()
    if tag == "Saturday":
        wb.active = 5
        sheet = wb.active
        for i in range(1, sheet.max_column + 1):
            print(sheet.cell(row=1, column=i).value, end='\t')
        print('\n')
        for r in range(2, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                print(sheet.cell(row=r, column=c).value, end='\t\t')
            print()
    if tag == "Sunday":
        wb.active = 6
        sheet = wb.active
        for i in range(1, sheet.max_column + 1):
            print(sheet.cell(row=1, column=i).value, end='\t')
        print('\n')
        for r in range(2, sheet.max_row + 1):
            for c in range(1, sheet.max_column + 1):
                print(sheet.cell(row=r, column=c).value, end='\t\t')
            print()
def funar(hoja, numrenglon):
    hoja.delete_rows(numrenglon+1, 1)
def renumber(ws):
    path = "POST IT APP.xlsx"
    wb = openpyxl.load_workbook(path)
    number = 0
    for r in range(2, ws.max_row + 1): #reenumera los elementos de la primer columna para que el usuario no confunda el numero de tareas
        number = number + 1
        for c in range(1, ws.max_column):
            ws.cell(row=r, column=c).value = number
    wb.save(".\POST IT APP.xlsx")
def dia(ws):
    #simplemente muestra el día de la hoja seleccionada
    for i in range(1, ws.max_column + 1):
        print(ws.cell(row=1, column=i).value, end='\t')
    print('\n')
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            print(ws.cell(row=r, column=c).value, end='\t\t')
        print()
def editarea(ws, d):
    path = "POST IT APP.xlsx"
    wb = openpyxl.load_workbook(path)
    m=input(colored("Modifica la tarea:", "red")) #la tarea a modificar
    while m == " ":
        m = input(colored("Por favor escribe algo", "red"))
    ws.cell(row = d+1, column = 2).value = m #la variable "d" es el numero de la tarea dada por el usuario y como solo contamos con dos columnas modificamos el valor del renglón deseado
    try:
        wb.save(".\POST IT APP.xlsx")
        print(colored("Modificado Exitosamente", "red"))
    except:
        print(colored("Ocurrió un error al escribir", "red"))
    finally:
        wb.close()
def showtasks(ws):
    #muestra la hoja activa sin headers
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            print(ws.cell(row=r, column=c).value, end='\t\t')
        print()
def pedirNumero():
    #pide el numero, sigue protocolos y parametros para verificar que la entrada sea correcta
    correcto = False
    num = 0
    while (not correcto):
        try:
            num = int(input("Introduce un numero entero: "))
            correcto = True
        except ValueError:
            print('Error, introduce un numero entero')
    return num
salir = False
opcion = 0
from colorama import init
from termcolor import colored
init()
day = (datetime.today().strftime('%A')) #día de la semana en la que el usuario abre la aplicación
menu(day)
while not salir:
    #menu principal
    wb = openpyxl.load_workbook('POST IT APP.xlsx')
    print(colored("1. Lunes", 'green'))
    print(colored("2. Martes", 'green'))
    print(colored("3. Miercoles", 'green'))
    print(colored("4. Jueves", 'green'))
    print(colored("5. Viernes", 'green'))
    print(colored("6. Sabado", "green"))
    print(colored("7. Domingo", "green"))
    print("Elige una opcion del 1 al 7, dependiendo del dia que busques. Escribe 0 para salir del programa")
    eleccion = int(input())
    while eleccion < 0 or eleccion > 7:
        eleccion = int(input("NO VÁLIDO Escribe el numero de elección de nuevo\n"))
    if eleccion == 0:
        print("Ten un día productivo, hasta luego")
        wb.close()
        exit(code=-1)
    wb.active = eleccion - 1
    sheet = wb.active
    renumber(sheet)
    dia(sheet)
    numcell = sheet['A2']
    taskscell = sheet['B2']
    salir = False
    opcion = 0
    print(" ")
    while not salir:
        #submenú
        print(colored("1. Modificar tarea", "red"))
        print(colored("2. Agregar tarea", "blue"))
        print(colored("3. Eliminar tarea", "magenta"))
        print(colored("4. Regresar al menu principal", "yellow"))
        print("Elige una opcion del 1 al 4")
        opcion = pedirNumero()

        if opcion == 1:
            dia(sheet)
            edit = int(input(colored("Escribe el numero de la tarea que deseas editar si deseas salir escribe un numero negativo o 0", "red")))
            if edit == 0:
                opcion = 4
                break
            editarea(sheet, edit)
            while edit > 0:
                edit = int(input(colored("Escribe el numero de la tarea que deseas editar si deseas salir escribe un numero negativo o 0", "red")))
                if edit <= 0:
                    dia(sheet)
                    wb.save(".\POST IT APP.xlsx")
                    opcion = 4
                    break
                while edit > sheet.max_row:
                    edit = int(input(colored("Reescribe el número", "red")))
                editarea(sheet, edit)
                dia(sheet)
        if opcion == 2:
            addtasks(sheet)
            renumber(sheet)
            showtasks(sheet)
            wb.save(".\POST IT APP.xlsx")
            wb.close()
        if opcion == 3:
            if ((sheet['B2'].value == None) and (sheet['A2'].value == None)):
                print(colored("¡No hay nada que borrar!", "magenta"))
                break
            dia(sheet)
            erasetarea = int(input(colored("Escribe el numero de la tarea que deseas borrar para salir escribe 0", "magenta")))
            if erasetarea == 0:
                break
            funar(sheet, erasetarea)
            wb.save(".\POST IT APP.xlsx")
            renumber(sheet)
            dia(sheet)
            if ((sheet['B2'].value == None) and (sheet['A2'].value == None)):
                print(colored("¡No hay nada que borrar!", "magenta"))
                break
            mercyless = str(input(colored("¿Quieres seguir borrando tareas?", "magenta")))
            while (mercyless == "Si" or mercyless == "SI" or mercyless == "si"):
                if ((sheet['B2'].value == None) and (sheet['A2'].value == None)):
                    print(colored("¡No hay nada que borrar!", "magenta"))
                    opcion = 4
                    break
                erasetarea = int(input(colored("Escribe el numero de la tarea que deseas borrar para salir escribe 0", "magenta")))
                if erasetarea == 0:
                    wb.save(".\POST IT APP.xlsx")
                    opcion = 4
                    break
                funar(sheet, erasetarea)
                renumber(sheet)
                showtasks(sheet)
                wb.save(".\POST IT APP.xlsx")

        if opcion == 4:
            break


